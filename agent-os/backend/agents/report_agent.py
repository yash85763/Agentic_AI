"""
ReportAgent - Assembles interactive narrative reports with sections,
charts, and an executive summary. Also exports to Excel via openpyxl.
"""

from __future__ import annotations

import io
import json
import logging
import os
import textwrap
from datetime import datetime
from typing import Any, Dict, List, Optional

from langfuse.decorators import observe

from agents.base_agent import BaseAgent

logger = logging.getLogger(__name__)

REPORT_MODEL = os.getenv("REPORT_MODEL", "claude-opus-4-5")
REPORT_FALLBACK = os.getenv("REPORT_FALLBACK", "openai/gpt-4o")


class ReportAgent(BaseAgent):
    """Generates the final narrative report and Excel export."""

    DEFAULT_MODEL: str = REPORT_MODEL

    DEFAULT_SECTIONS = (
        "overview",
        "key_findings",
        "team_breakdown",
        "anomalies",
        "recommendations",
    )

    def __init__(
        self,
        cognitive_context: Dict[str, Any] = None,
        langfuse_handler: Any = None,
        model: str = None,
    ) -> None:
        super().__init__(
            model=model or REPORT_MODEL,
            cognitive_context=cognitive_context or {},
            langfuse_handler=langfuse_handler,
        )

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    @observe(name="report.run")
    def run(self, state: Dict[str, Any]) -> Dict[str, Any]:
        """Assemble the final report.

        Expected state:
            - validation: validation report dict
            - charts: list of chart configs
            - data_summary: dict with aggregate stats
            - task: original task description
            - job_id, redis_client, merged_path
        """
        job_id = state.get("job_id", self._job_id())
        redis_client = state.get("redis_client")

        validation = state.get("validation", {})
        charts = state.get("charts", [])
        data_summary = state.get("data_summary", {})
        task = state.get("task", "Data analysis")

        self._emit_event(
            "agent_start",
            {"agent": "report"},
            job_id,
            redis_client,
        )

        # 1. Executive summary
        exec_summary = self.generate_executive_summary(
            task=task,
            validation=validation,
            data_summary=data_summary,
            chart_count=len(charts),
        )
        self._emit_event(
            "report_section",
            {"section": "executive_summary", "content": exec_summary},
            job_id,
            redis_client,
        )

        # 2. Generate each section
        sections: List[Dict[str, Any]] = []
        for section_type in self.DEFAULT_SECTIONS:
            try:
                section = self.generate_section(
                    section_type=section_type,
                    validation=validation,
                    charts=charts,
                    data_summary=data_summary,
                    task=task,
                )
                sections.append(section)
                self._emit_event(
                    "report_section",
                    {"section": section_type, "content": section["content"]},
                    job_id,
                    redis_client,
                )
            except Exception as exc:
                logger.warning("Failed to generate section '%s': %s", section_type, exc)

        # 3. Assemble final report
        report = self.assemble_report(
            task=task,
            executive_summary=exec_summary,
            sections=sections,
            charts=charts,
            validation=validation,
        )
        state["report"] = report

        # 4. Excel export (best-effort)
        merged_path = state.get("merged_path")
        if merged_path and os.path.exists(merged_path):
            try:
                excel_bytes = self.export_excel(report, merged_path, charts)
                state["excel_bytes"] = excel_bytes
                state["excel_size"] = len(excel_bytes)
            except Exception as exc:
                logger.warning("Excel export failed: %s", exc)

        self._emit_event(
            "agent_complete",
            {"agent": "report", "section_count": len(sections)},
            job_id,
            redis_client,
        )
        self._emit_event(
            "complete",
            {"report_id": report.get("job_id"), "section_count": len(sections), "chart_count": len(charts)},
            job_id,
            redis_client,
        )
        return state

    # ------------------------------------------------------------------
    # Section generation
    # ------------------------------------------------------------------

    def generate_executive_summary(
        self,
        task: str,
        validation: Dict[str, Any],
        data_summary: Dict[str, Any],
        chart_count: int,
    ) -> str:
        """Generate an executive summary using the LLM."""
        facts = self._build_facts_block(validation, data_summary, chart_count)

        system = textwrap.dedent(
            """
            You are the Report Agent of an open-source agentic data platform.

            Write a tight executive summary (3-5 sentences) of the analysis.

            STRICT RULES:
            - Only use the facts and numbers provided below — never invent or
              embellish numbers
            - Tone: precise, business-friendly, no fluff
            - Format: plain markdown, no headings (this is the summary block)
            - Cite specific numbers from the FACTS section
            - If validation failed, lead with that
            """
        ).strip()

        user = f"Task: {task}\n\nFACTS:\n{facts}\n\nWrite the executive summary."

        response = self._call_llm(
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            temperature=0.3,
            max_tokens=800,
            fallback_model=REPORT_FALLBACK,
        )
        return self._extract_text(response).strip()

    def generate_section(
        self,
        section_type: str,
        validation: Dict[str, Any],
        charts: List[Dict[str, Any]],
        data_summary: Dict[str, Any],
        task: str,
    ) -> Dict[str, Any]:
        """Generate a single report section."""
        prompts = {
            "overview": "Provide a 2-3 paragraph overview of the dataset: row count, time range covered, scope.",
            "key_findings": "Identify 3-5 key findings as a bullet list with specific numbers.",
            "team_breakdown": "Break down results by team/category. Use a markdown table.",
            "anomalies": "List any anomalies, outliers, or data-quality issues flagged by validation.",
            "recommendations": "Provide 3-5 actionable recommendations based on the findings.",
        }

        prompt_body = prompts.get(section_type, f"Generate the {section_type} section.")
        facts = self._build_facts_block(validation, data_summary, len(charts))

        system = textwrap.dedent(
            f"""
            You are the Report Agent. Write the "{section_type}" section in markdown.

            STRICT RULES:
            - Only use the FACTS provided — never invent numbers
            - Use proper markdown: ## headings, **bold**, - bullets, | tables
            - Keep it focused and data-driven
            - If a section can't be supported by the facts, write "No data available" rather than speculating

            Task context: {task}
            """
        ).strip()

        user = f"FACTS:\n{facts}\n\nSection instruction: {prompt_body}"

        response = self._call_llm(
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            temperature=0.3,
            max_tokens=1500,
            fallback_model=REPORT_FALLBACK,
        )

        content = self._extract_text(response).strip()
        return {
            "id": section_type,
            "title": section_type.replace("_", " ").title(),
            "content": content,
            "chart": None,
        }

    # ------------------------------------------------------------------
    # Report assembly
    # ------------------------------------------------------------------

    def assemble_report(
        self,
        task: str,
        executive_summary: str,
        sections: List[Dict[str, Any]],
        charts: List[Dict[str, Any]],
        validation: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Combine all components into the final report dict."""
        # Attach charts to relevant sections
        if charts:
            # Distribute charts: first chart to overview, rest to key_findings
            chart_map = {
                "overview": charts[0] if charts else None,
                "key_findings": charts[1] if len(charts) > 1 else None,
                "team_breakdown": charts[2] if len(charts) > 2 else None,
            }
            for section in sections:
                if section["id"] in chart_map and chart_map[section["id"]]:
                    section["chart"] = chart_map[section["id"]]

        return {
            "job_id": self._job_id(),
            "title": self._generate_title(task),
            "task": task,
            "executive_summary": executive_summary,
            "sections": sections,
            "charts": charts,
            "validation_summary": {
                "passed": validation.get("passed", False),
                "row_count": validation.get("row_count", 0),
                "anomaly_count": len(validation.get("anomalies", [])),
                "check_count": len(validation.get("checks", [])),
            },
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "version": "2.0",
        }

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def export_excel(
        self,
        report: Dict[str, Any],
        merged_path: str,
        charts: List[Dict[str, Any]],
    ) -> bytes:
        """Export the report to an Excel file with embedded data and charts."""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        except ImportError as exc:
            logger.error("openpyxl or pandas not available for Excel export: %s", exc)
            return b""

        wb = Workbook()

        # === Sheet 1: Executive Summary ===
        ws = wb.active
        ws.title = "Summary"

        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

        ws["A1"] = report.get("title", "Report")
        ws["A1"].font = Font(bold=True, size=18)
        ws.merge_cells("A1:E1")

        ws["A3"] = "Generated"
        ws["B3"] = report.get("generated_at", "")
        ws["A4"] = "Task"
        ws["B4"] = report.get("task", "")
        ws.merge_cells("B4:E4")
        ws["B4"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[4].height = 60

        ws["A6"] = "Executive Summary"
        ws["A6"].font = header_font
        ws["A6"].fill = header_fill
        ws.merge_cells("A6:E6")

        ws["A7"] = report.get("executive_summary", "")
        ws.merge_cells("A7:E15")
        ws["A7"].alignment = Alignment(wrap_text=True, vertical="top")

        # Validation summary
        vs = report.get("validation_summary", {})
        ws["A17"] = "Validation"
        ws["A17"].font = header_font
        ws["A17"].fill = header_fill
        ws.merge_cells("A17:E17")

        ws["A18"] = "Passed"
        ws["B18"] = "✓ Yes" if vs.get("passed") else "✗ No"
        ws["A19"] = "Row Count"
        ws["B19"] = vs.get("row_count", 0)
        ws["A20"] = "Anomalies"
        ws["B20"] = vs.get("anomaly_count", 0)

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 25
        for col in ("C", "D", "E"):
            ws.column_dimensions[col].width = 18

        # === Sheet 2: Sections ===
        sw = wb.create_sheet("Sections")
        row = 1
        for section in report.get("sections", []):
            sw.cell(row=row, column=1, value=section.get("title", ""))
            sw.cell(row=row, column=1).font = header_font
            sw.cell(row=row, column=1).fill = header_fill
            sw.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1
            sw.cell(row=row, column=1, value=section.get("content", ""))
            sw.merge_cells(start_row=row, start_column=1, end_row=row + 6, end_column=5)
            sw.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            row += 8

        sw.column_dimensions["A"].width = 20
        for col in ("B", "C", "D", "E"):
            sw.column_dimensions[col].width = 18

        # === Sheet 3: Raw Data ===
        try:
            df = pd.read_parquet(merged_path)
            dw = wb.create_sheet("Data")
            for r in dataframe_to_rows(df, index=False, header=True):
                dw.append(r)
            # Style header row
            for cell in dw[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        except Exception as exc:
            logger.warning("Could not embed data sheet: %s", exc)

        # === Sheet 4: Charts metadata ===
        if charts:
            cw = wb.create_sheet("Charts")
            cw["A1"] = "Chart"
            cw["B1"] = "Type"
            cw["C1"] = "Config (JSON)"
            for cell in (cw["A1"], cw["B1"], cw["C1"]):
                cell.font = header_font
                cell.fill = header_fill

            for i, chart in enumerate(charts, start=2):
                cw.cell(row=i, column=1, value=chart.get("title", f"Chart {i-1}"))
                cw.cell(row=i, column=2, value=chart.get("type", ""))
                cw.cell(row=i, column=3, value=json.dumps(chart.get("config", {}))[:32000])

            cw.column_dimensions["A"].width = 30
            cw.column_dimensions["B"].width = 15
            cw.column_dimensions["C"].width = 80

        # Render to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _build_facts_block(
        self,
        validation: Dict[str, Any],
        data_summary: Dict[str, Any],
        chart_count: int,
    ) -> str:
        """Compose a precise, structured FACTS block for the LLM."""
        parts: List[str] = []

        parts.append(f"VALIDATION PASSED: {validation.get('passed', False)}")
        parts.append(f"ROW COUNT: {validation.get('row_count', 0)}")
        parts.append(f"TOTAL VALUE: {validation.get('total_value', 0)}")
        parts.append(f"CHART COUNT: {chart_count}")

        # Checks
        checks = validation.get("checks", [])
        if checks:
            parts.append("\nVALIDATION CHECKS:")
            for check in checks[:10]:
                parts.append(
                    f"  - {check.get('name')}: "
                    f"{'PASS' if check.get('passed') else 'FAIL'} — "
                    f"{check.get('message', '')}"
                )

        # Anomalies
        anomalies = validation.get("anomalies", [])
        if anomalies:
            parts.append(f"\nANOMALIES ({len(anomalies)} found):")
            for a in anomalies[:5]:
                parts.append(
                    f"  - {a.get('severity', 'info').upper()}: "
                    f"{a.get('column', '')} — {a.get('description', '')}"
                )

        # Group breakdowns
        groups = data_summary.get("groups", {})
        if groups:
            parts.append("\nGROUP TOTALS:")
            for group_col, values in list(groups.items())[:3]:
                parts.append(f"  By {group_col}:")
                for k, v in list(values.items())[:8]:
                    parts.append(f"    - {k}: {v}")

        return "\n".join(parts)

    def _generate_title(self, task: str) -> str:
        if len(task) <= 80:
            return task
        return task[:77] + "..."
